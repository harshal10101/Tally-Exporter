"""
Excel Generator Service
Generates Tally-compatible Excel file with 24 columns
"""

import io
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# Tally export columns (24 columns - final structure with Product)
TALLY_COLUMNS = [
    "Sr. No.",
    "Invoice Type",
    "Product",
    "Invoice No",
    "Invoice Date",
    "GST Registration",
    "GST State",
    "Party/Customer",
    "Order No",
    "Order Date",
    "Invoice Period From",
    "Invoice Period To",
    "Billing Frequency",
    "TDS Applicable",
    "GST TDS Applicable",
    "Ledger Name",
    "Submitted Qty",
    "Submitted Rate",
    "Delivered Qty",
    "Delivered Rate",
    "Amount",
    "CGST",
    "SGST",
    "Total Amount (with Tax)"
]


def generate_excel(data: List[Dict]) -> io.BytesIO:
    """
    Generate Excel file with extracted invoice data in Tally format.
    
    Args:
        data: List of extracted invoice data dictionaries
        
    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tally Import"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_idx, header in enumerate(TALLY_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_idx, invoice in enumerate(data, 2):
        # Format invoice_type for display (uppercase)
        invoice_type = invoice.get("invoice_type", "").upper()
        
        row_data = [
            row_idx - 1,  # Sr. No.
            invoice_type,  # Invoice Type (CloudXP/JTL/RJIL)
            invoice.get("product", "SMS"),  # Product (default: SMS)
            invoice.get("invoice_no", ""),
            invoice.get("invoice_date", ""),
            invoice.get("gst_registration", ""),
            invoice.get("gst_state", ""),
            invoice.get("party_customer", ""),
            invoice.get("order_no", ""),
            invoice.get("order_date", ""),
            invoice.get("invoice_period_from", ""),
            invoice.get("invoice_period_to", ""),
            invoice.get("billing_frequency", ""),
            invoice.get("tds_applicable", "Yes"),
            invoice.get("gst_tds_applicable", "No"),
            invoice.get("ledger_name", ""),
            invoice.get("submitted_qty", ""),
            invoice.get("submitted_rate", ""),
            invoice.get("delivered_qty", ""),
            invoice.get("delivered_rate", ""),
            invoice.get("amount", ""),
            invoice.get("cgst", ""),
            invoice.get("sgst", ""),
            invoice.get("total_amount", "")
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # Auto-adjust column widths
    for col_idx in range(1, len(TALLY_COLUMNS) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = len(TALLY_COLUMNS[col_idx - 1])
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
