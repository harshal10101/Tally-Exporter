"""
Unit tests for Excel and CSV generator services.
"""

import pytest
import io
from openpyxl import load_workbook
from services.excel_generator import generate_excel, TALLY_COLUMNS
from services.csv_generator import generate_csv, TALLY_COLUMNS as CSV_COLUMNS


SAMPLE_DATA = [
    {
        "invoice_type": "cloudxp",
        "product": "SMS",
        "invoice_no": "INV001",
        "invoice_date": "15/11/2025",
        "gst_registration": "27AABCN1234Q1ZM",
        "gst_state": "Maharashtra",
        "party_customer": "TEST COMPANY",
        "order_no": "PO001",
        "order_date": "01/11/2025",
        "invoice_period_from": "01/11/2025",
        "invoice_period_to": "30/11/2025",
        "billing_frequency": "Monthly",
        "tds_applicable": "Yes",
        "gst_tds_applicable": "No",
        "ledger_name": "Bulk SMS Charges - Nov-25 to Nov-25",
        "submitted_qty": "100000",
        "submitted_rate": "0.020000",
        "delivered_qty": "90000",
        "delivered_rate": "0.090000",
        "amount": "10000.00",
        "cgst": "900.00",
        "sgst": "900.00",
        "total_amount": "11800.00",
    }
]


class TestExcelGenerator:
    def test_generates_valid_excel(self):
        buffer = generate_excel(SAMPLE_DATA)
        assert buffer is not None
        assert buffer.getvalue()[:4] == b"PK\x03\x04"  # ZIP/XLSX magic bytes

    def test_has_correct_headers(self):
        buffer = generate_excel(SAMPLE_DATA)
        wb = load_workbook(io.BytesIO(buffer.getvalue()))
        ws = wb.active
        headers = [ws.cell(row=1, column=i).value for i in range(1, len(TALLY_COLUMNS) + 1)]
        assert headers == TALLY_COLUMNS

    def test_has_data_row(self):
        buffer = generate_excel(SAMPLE_DATA)
        wb = load_workbook(io.BytesIO(buffer.getvalue()))
        ws = wb.active
        assert ws.cell(row=2, column=4).value == "INV001"

    def test_empty_data(self):
        buffer = generate_excel([])
        wb = load_workbook(io.BytesIO(buffer.getvalue()))
        ws = wb.active
        assert ws.max_row == 1  # Only headers

    def test_sheet_name(self):
        buffer = generate_excel(SAMPLE_DATA)
        wb = load_workbook(io.BytesIO(buffer.getvalue()))
        assert wb.active.title == "Tally Import"


class TestCSVGenerator:
    def test_generates_valid_csv(self):
        buffer = generate_csv(SAMPLE_DATA)
        content = buffer.getvalue().decode("utf-8-sig")
        assert "Sr. No." in content
        assert "INV001" in content

    def test_has_correct_headers(self):
        buffer = generate_csv(SAMPLE_DATA)
        content = buffer.getvalue().decode("utf-8-sig")
        first_line = content.split("\r\n")[0]
        for col in ["Sr. No.", "Invoice Type", "Invoice No", "Total Amount (with Tax)"]:
            assert col in first_line

    def test_has_data_row(self):
        buffer = generate_csv(SAMPLE_DATA)
        content = buffer.getvalue().decode("utf-8-sig")
        lines = content.strip().split("\r\n")
        assert len(lines) == 2  # header + 1 data row
        assert "INV001" in lines[1]

    def test_empty_data(self):
        buffer = generate_csv([])
        content = buffer.getvalue().decode("utf-8-sig")
        lines = content.strip().split("\r\n")
        assert len(lines) == 1  # header only
